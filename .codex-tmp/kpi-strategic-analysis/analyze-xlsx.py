import json
import re
from collections import Counter, defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


BOOK_PATH = Path(r"C:\Users\rguimaraes\Downloads\#INDICADORES# 2026.xlsx")

wb_formula = load_workbook(BOOK_PATH, data_only=False, read_only=False)
wb_values = load_workbook(BOOK_PATH, data_only=True, read_only=False)

sheet_report = []
cross_refs = Counter()
broken_formulas = []
external_formulas = []

for ws in wb_formula.worksheets:
    value_ws = wb_values[ws.title]
    formulas = []
    constants = []
    formula_functions = Counter()
    formula_cells_by_col = Counter()
    constant_cells_by_col = Counter()
    errors = []
    comments = []
    hyperlinks = []
    cross_sheet_refs = Counter()

    for row in ws.iter_rows():
        for cell in row:
            value = cell.value
            if value is None:
                continue
            if isinstance(value, str) and value.startswith("="):
                formulas.append(cell.coordinate)
                formula_cells_by_col[cell.column] += 1
                for fn in re.findall(r"\b([A-Z][A-Z0-9\.]*)\s*\(", value.upper()):
                    formula_functions[fn] += 1
                for ref in re.findall(r"(?:'([^']+)'|([A-Za-z0-9_ À-ÿ]+))!\$?[A-Z]{1,3}\$?\d+", value):
                    ref_name = (ref[0] or ref[1]).strip()
                    if ref_name:
                        cross_sheet_refs[ref_name] += 1
                        cross_refs[(ws.title, ref_name)] += 1
                if "#REF!" in value:
                    broken_formulas.append({"sheet": ws.title, "cell": cell.coordinate, "formula": value})
                if "[" in value and "]" in value:
                    external_formulas.append({"sheet": ws.title, "cell": cell.coordinate, "formula": value})
            else:
                constants.append(cell.coordinate)
                constant_cells_by_col[cell.column] += 1
            cached = value_ws[cell.coordinate].value
            if isinstance(cached, str) and cached.startswith("#"):
                errors.append({"cell": cell.coordinate, "value": cached, "formula": value if isinstance(value, str) and value.startswith("=") else None})
            if cell.comment:
                comments.append({"cell": cell.coordinate, "author": cell.comment.author, "text": cell.comment.text[:500]})
            if cell.hyperlink:
                hyperlinks.append({"cell": cell.coordinate, "target": cell.hyperlink.target or cell.hyperlink.location})

    # Detect source-table KPI blocks. The workbook consistently stores chart sources from column V onward.
    row1_headers = []
    for col in range(22, ws.max_column + 1):
        cell = ws.cell(1, col)
        if cell.value not in (None, ""):
            row1_headers.append({
                "cell": cell.coordinate,
                "text": str(cell.value).strip(),
                "next_headers": [
                    ws.cell(1, c).value for c in range(col + 1, min(ws.max_column, col + 6) + 1)
                ],
            })

    kpi_starts = []
    for item in row1_headers:
        col = ws[item["cell"]].column
        nearby = [str(ws.cell(1, c).value or "").strip().lower() for c in range(col + 1, min(ws.max_column, col + 5) + 1)]
        if any(v in {"mês", "mes", "meses"} for v in nearby):
            month_col_offset = next(i for i, v in enumerate(nearby, start=1) if v in {"mês", "mes", "meses"})
            month_col = col + month_col_offset
            months = [value_ws.cell(r, month_col).value for r in range(2, min(13, ws.max_row) + 1)]
            kpi_starts.append({
                "title_cell": item["cell"],
                "title": item["text"],
                "month_col": get_column_letter(month_col),
                "months": [str(v) if v is not None else None for v in months],
                "fields": [str(ws.cell(1, c).value).strip() if ws.cell(1, c).value is not None else None for c in range(col, min(ws.max_column, col + 13) + 1)],
            })

    # Refine to true KPI starts: the title is immediately followed by Mês/Meses.
    true_start_cols = []
    for col in range(22, ws.max_column + 1):
        current = ws.cell(1, col).value
        next_value = str(ws.cell(1, col + 1).value or "").strip().lower() if col < ws.max_column else ""
        if current not in (None, "") and next_value in {"mês", "mes", "meses"}:
            true_start_cols.append(col)

    kpi_blocks = []
    for start_col in true_start_cols:
        # KPI fields are contiguous in row 1; stop at the first blank header.
        end_col = start_col
        while end_col + 1 <= ws.max_column and ws.cell(1, end_col + 1).value not in (None, ""):
            end_col += 1
        fields = []
        for col in range(start_col, end_col + 1):
            header = str(ws.cell(1, col).value or "").strip()
            month_cells = [ws.cell(row, col) for row in range(2, min(13, ws.max_row) + 1)]
            formula_samples = []
            constants_in_period = []
            blank_count = 0
            for cell in month_cells:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    if cell.value not in formula_samples and len(formula_samples) < 4:
                        formula_samples.append(cell.value)
                elif cell.value is None:
                    blank_count += 1
                else:
                    constants_in_period.append(value_ws[cell.coordinate].value)
            fields.append({
                "column": get_column_letter(col),
                "header": header,
                "formula_count": sum(1 for cell in month_cells if isinstance(cell.value, str) and cell.value.startswith("=")),
                "constant_count": len(constants_in_period),
                "blank_count": blank_count,
                "constant_values": constants_in_period[:12],
                "formula_samples": formula_samples,
                "summary_cell": ws.cell(14, col).coordinate if ws.max_row >= 14 else None,
                "summary_formula": ws.cell(14, col).value if ws.max_row >= 14 and isinstance(ws.cell(14, col).value, str) and ws.cell(14, col).value.startswith("=") else None,
                "summary_value": value_ws.cell(14, col).value if ws.max_row >= 14 else None,
            })

        historical = []
        for row in range(15, min(22, ws.max_row) + 1):
            vals = []
            for col in range(start_col, end_col + 1):
                val = value_ws.cell(row, col).value
                if val is not None:
                    vals.append({"cell": ws.cell(row, col).coordinate, "value": val})
            if vals:
                historical.append({"row": row, "values": vals})

        within_formulas = []
        for field in fields:
            if "dentro" in field["header"].lower() or "fora" in field["header"].lower():
                within_formulas.extend(field["formula_samples"])
                if field["summary_formula"]:
                    within_formulas.append(field["summary_formula"])
        comparison_tokens = Counter()
        for formula in within_formulas:
            for token in re.findall(r">=|<=|<>|>|<", formula):
                comparison_tokens[token] += 1

        kpi_blocks.append({
            "title": str(ws.cell(1, start_col).value).strip(),
            "range": f"{get_column_letter(start_col)}1:{get_column_letter(end_col)}14",
            "fields": fields,
            "comparison_tokens": comparison_tokens.most_common(),
            "historical": historical,
        })

    hidden_rows = [idx for idx, dim in ws.row_dimensions.items() if dim.hidden]
    hidden_cols = [idx for idx, dim in ws.column_dimensions.items() if dim.hidden]
    sheet_report.append({
        "name": ws.title,
        "state": ws.sheet_state,
        "max_row": ws.max_row,
        "max_col": ws.max_column,
        "dimensions": ws.calculate_dimension(),
        "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
        "sheet_view": {
            "show_grid_lines": ws.sheet_view.showGridLines,
            "zoom_scale": ws.sheet_view.zoomScale,
        },
        "print_area": str(ws.print_area) if ws.print_area else None,
        "print_titles": str(ws.print_titles) if ws.print_titles else None,
        "hidden_rows": hidden_rows,
        "hidden_cols": hidden_cols,
        "merged_ranges_count": len(ws.merged_cells.ranges),
        "merged_ranges_sample": [str(rng) for rng in list(ws.merged_cells.ranges)[:30]],
        "tables": [{"name": table.name, "ref": table.ref} for table in ws.tables.values()],
        "charts": len(ws._charts),
        "images": len(ws._images),
        "data_validations": [{"sqref": str(dv.sqref), "type": dv.type, "formula1": dv.formula1, "formula2": dv.formula2} for dv in ws.data_validations.dataValidation],
        "protection": {"sheet": ws.protection.sheet, "selectLockedCells": ws.protection.selectLockedCells, "selectUnlockedCells": ws.protection.selectUnlockedCells},
        "formula_count": len(formulas),
        "constant_count": len(constants),
        "formula_functions": formula_functions.most_common(20),
        "formula_columns": [{"col": get_column_letter(k), "count": v} for k, v in formula_cells_by_col.most_common()],
        "constant_columns": [{"col": get_column_letter(k), "count": v} for k, v in constant_cells_by_col.most_common()],
        "cached_errors": errors[:100],
        "comments": comments,
        "hyperlinks": hyperlinks,
        "cross_sheet_refs": cross_sheet_refs.most_common(),
        "row1_headers": row1_headers,
        "kpi_starts": kpi_starts,
        "kpi_blocks": kpi_blocks,
    })

defined_names = []
for name, item in wb_formula.defined_names.items():
    defined_names.append({"name": name, "attr_text": item.attr_text, "hidden": item.hidden})

report = {
    "workbook": {
        "sheet_count": len(wb_formula.sheetnames),
        "sheetnames": wb_formula.sheetnames,
        "defined_names": defined_names,
        "calculation": {
            "calc_mode": wb_formula.calculation.calcMode,
            "full_calc_on_load": wb_formula.calculation.fullCalcOnLoad,
            "force_full_calc": wb_formula.calculation.forceFullCalc,
        },
        "vba_archive": wb_formula.vba_archive is not None,
    },
    "sheets": sheet_report,
    "cross_sheet_edges": [{"from": a, "to": b, "count": n} for (a, b), n in cross_refs.most_common()],
    "broken_formulas": broken_formulas,
    "external_formulas": external_formulas,
}

Path("openpyxl-report.json").write_text(json.dumps(report, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
print(json.dumps({
    "sheets": [{"name": s["name"], "kpis": len(s["kpi_starts"]), "formulas": s["formula_count"], "charts": s["charts"], "errors": len(s["cached_errors"]), "hidden_cols": len(s["hidden_cols"])} for s in sheet_report],
    "broken_formulas": len(broken_formulas),
    "external_formulas": len(external_formulas),
    "cross_sheet_edges": report["cross_sheet_edges"],
}, ensure_ascii=False, indent=2))
